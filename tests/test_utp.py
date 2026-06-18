from pathlib import Path

from jac_scraper.utp import normalize_series, match_series, UtpCandidate


def test_normalize_series_uppercase_and_spaces():
    assert normalize_series("  Integra   pro  ") == "INTEGRA PRO"
    assert normalize_series("aurora on/off r32") == "AURORA ON/OFF R32"
    assert normalize_series("") == ""
    assert normalize_series(None) == ""


def test_match_series_direct_and_alias():
    candidate_index = {"INTEGRA PRO": "Integra Pro", "AURORA": "Aurora"}
    assert match_series("integra pro", candidate_index) == "Integra Pro"
    aliases = {"AURORA ON/OFF R32": "AURORA"}
    assert match_series("Aurora ON/OFF R32", candidate_index, aliases) == "Aurora"
    assert match_series("НЕИЗВЕСТНАЯ", candidate_index) is None


def test_write_candidates_xlsx_and_json(tmp_path: Path):
    from jac_scraper.utp import UtpCandidate, write_candidates

    cands = [
        UtpCandidate(brand="MDV", series="INTEGRA", text="3D Air Flow"),
        UtpCandidate(brand="MDV", series="INTEGRA", text="Wi-Fi управление"),
        UtpCandidate(brand="THAICON", series="PHANTOM", text="Тихий режим"),
    ]
    xlsx_path = tmp_path / "jac_utp_candidates.xlsx"
    json_path = tmp_path / "jac_utp_candidates.json"
    write_candidates(cands, xlsx_path, json_path)

    assert xlsx_path.exists() and json_path.exists()

    from openpyxl import load_workbook
    ws = load_workbook(xlsx_path).active
    header = [c.value for c in ws[1]]
    assert header == ["Бренд", "Серия", "№", "Текст УТП", "Брать"]
    assert [ws.cell(row=2, column=i).value for i in range(1, 5)] == ["MDV", "INTEGRA", 1, "3D Air Flow"]
    assert ws.cell(row=2, column=5).value is None


from jac_scraper.utp import build_latest_from_xlsx


def test_build_latest_only_marked_rows(tmp_path: Path):
    from openpyxl import Workbook
    from jac_scraper.utp import CANDIDATES_HEADER
    wb = Workbook(); ws = wb.active
    ws.append(CANDIDATES_HEADER)
    ws.append(["MDV", "Integra", 1, "3D Air Flow", "x"])
    ws.append(["MDV", "Integra", 2, "Мусорный пункт", None])
    ws.append(["MDV", "Aurora", 1, "Компактный корпус", "1"])
    xlsx_path = tmp_path / "in.xlsx"; wb.save(xlsx_path)

    out_path = tmp_path / "jac_utp_latest.json"
    result = build_latest_from_xlsx(xlsx_path, out_path)

    assert result == {
        "MDV": {"INTEGRA": ["3D Air Flow"], "AURORA": ["Компактный корпус"]}
    }
    import json as _j
    assert _j.loads(out_path.read_text(encoding="utf-8")) == result


from jac_scraper.utp import replace_brand_candidates, write_candidates


def test_replace_brand_candidates_preserves_other_brand_ticks(tmp_path: Path):
    from openpyxl import load_workbook
    xlsx_path = tmp_path / "jac_utp_candidates.xlsx"
    json_path = tmp_path / "jac_utp_candidates.json"
    # исходный файл: MDV (старое) + THAICON с галочкой
    write_candidates([
        UtpCandidate("MDV", "OLD", "старое MDV"),
        UtpCandidate("THAICON", "PHANTOM", "Тихий режим"),
    ], xlsx_path, json_path)
    ws = load_workbook(xlsx_path).active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "THAICON":
            row[4].value = "x"            # пользователь отметил THAICON
    ws.parent.save(xlsx_path)

    # заменяем MDV новыми строками
    kept = replace_brand_candidates(xlsx_path, json_path, "MDV",
                                    [UtpCandidate("MDV", "INTEGRA", "новое MDV")])
    assert kept == 1                       # одна чужая (THAICON) строка сохранена

    ws = load_workbook(xlsx_path).active
    rows = [(r[0], r[1], r[3], r[4]) for r in ws.iter_rows(min_row=2, values_only=True)]
    # старое MDV исчезло, новое появилось, галочка THAICON на месте
    assert ("MDV", "OLD", "старое MDV", None) not in rows
    assert ("MDV", "INTEGRA", "новое MDV", None) in rows
    assert ("THAICON", "PHANTOM", "Тихий режим", "x") in rows


from jac_scraper.utp import coverage_gaps


def test_coverage_gaps():
    # INTEGRA покрыта кандидатом; AURORA — пробел; «Сплит-система ...» — тип, не серия
    jac_series = {"MDV": ["INTEGRA", "Aurora", "Сплит-система настенного типа"]}
    candidates = [UtpCandidate("MDV", "Integra", "x")]
    type_words = ("сплит-система", "канальн", "кассетн", "мульти")
    gaps = coverage_gaps(jac_series, candidates, type_words)
    assert gaps == {"MDV": ["AURORA"]}
