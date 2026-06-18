"""УТП серий MDV из официального прайс-листа (xlsx), а не с сайта.

В прайсе на листах «RAC inverter» и «RAC on-off» каждая серия — это строка-заголовок
(колонка A, напр. «INTEGRA ON/OFF, R32»), а под ней высокая строка с УТП: маркеры «●»
разнесены по колонкам A–C, перенос строки = отдельный пункт. Источник полнее и точнее
сайта (там не было on/off-серий).

Имена серий в прайсе содержат маркетинговые хвосты («ERP Full DC INVERTER R32»),
поэтому сопоставляем с сериями JAC по вхождению токенов (все слова имени JAC есть в
имени прайса; выбираем самое «длинное» совпадение). Серии, которых нет в выгрузке JAC,
пропускаются (их всё равно некому показывать в боте).
"""
from __future__ import annotations

import re
from typing import List, Tuple

from openpyxl import load_workbook

from .utp import UtpCandidate, normalize_series

PRICELIST_SHEETS = ("RAC inverter", "RAC on-off")
_BULLET = "●"
_UTP_COLS = range(1, 7)  # A–F: в этих колонках лежат маркеры УТП


def clean_series_label(raw: str) -> str:
    """«INTEGRA ON/OFF, R32 (снято с производства)» -> «INTEGRA ON/OFF R32»."""
    s = re.sub(r"\(.*?\)", "", raw or "")   # убрать (снято с производства) и т.п.
    s = s.replace(",", " ")
    return " ".join(s.split())


def build_jac_index(products: List[dict], brand: str = "MDV") -> List[Tuple[set, str]]:
    """[(множество_токенов_имени, каноничное_имя_JAC)] для серий бренда из stock."""
    canon: dict = {}
    for p in products:
        if p.get("brand") == brand and (p.get("series") or "").strip():
            canon.setdefault(normalize_series(p["series"]), p["series"].strip())
    return [(set(norm.split()), name) for norm, name in canon.items()]


def map_to_jac(pl_series: str, jac_index: List[Tuple[set, str]]) -> str | None:
    """Имя серии из прайса -> каноничное имя серии JAC (или None).

    Совпадение: все токены имени JAC содержатся в токенах имени прайса. Из подходящих
    берём с наибольшим числом токенов (чтобы «INTEGRA PRO» побеждал «INTEGRA»).
    """
    tokens = set(normalize_series(pl_series).split())
    best = None  # (число_токенов, имя)
    for jac_tokens, name in jac_index:
        if jac_tokens <= tokens:
            if best is None or len(jac_tokens) > best[0]:
                best = (len(jac_tokens), name)
    return best[1] if best else None


def parse_mdv_pricelist(path, jac_index: List[Tuple[set, str]]
                        ) -> Tuple[List[UtpCandidate], List[str]]:
    """Парсит прайс MDV -> (кандидаты с именами серий JAC, список несопоставленных серий).

    Строка УТП определяется по наличию «●» в колонках A–F; имя серии берём из колонки A
    строки прямо над ней.
    """
    wb = load_workbook(path, data_only=True)
    cands: List[UtpCandidate] = []
    unmapped: List[str] = []
    for sheet in PRICELIST_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            bullets: List[str] = []
            for col in _UTP_COLS:
                v = ws.cell(row=r, column=col).value
                if v and _BULLET in str(v):
                    for line in str(v).split("\n"):
                        t = line.strip().lstrip(_BULLET).strip()
                        if t and t not in bullets:
                            bullets.append(t)
            if not bullets:
                continue
            series = clean_series_label(str(ws.cell(row=r - 1, column=1).value or ""))
            jac = map_to_jac(series, jac_index)
            if jac is None:
                if series and series not in unmapped:
                    unmapped.append(series)
                continue
            for t in bullets:
                cands.append(UtpCandidate(brand="MDV", series=jac, text=t))
    return cands, unmapped
