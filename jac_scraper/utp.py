"""УТП по сериям: каркас сбора кандидатов и сборки финального файла.

Поток (см. дизайн-док):
  utp-collect -> data/jac_utp_candidates.xlsx (+ .json бэкап)
  владелец ставит галочки в колонке «Брать»
  utp-build   -> data/jac_utp_latest.json {бренд: {СЕРИЯ_НОРМ: [УТП...]}}
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import List


def normalize_series(name: str) -> str:
    """Верхний регистр + схлопывание пробелов. Ключ для связки JAC <-> сайт."""
    return " ".join((name or "").upper().split())


def match_series(jac_series: str, candidate_index: dict, aliases: dict | None = None) -> str | None:
    """Возвращает оригинальное имя серии сайта для серии JAC или None.

    candidate_index: {НОРМ_имя_серии_сайта: оригинальное_имя_сайта}
    aliases: {НОРМ_имя_серии_JAC: НОРМ_имя_серии_сайта} — для расхождений написания
    """
    aliases = aliases or {}
    key = normalize_series(jac_series)
    key = aliases.get(key, key)
    return candidate_index.get(key)


CANDIDATES_XLSX = "jac_utp_candidates.xlsx"
CANDIDATES_JSON = "jac_utp_candidates.json"
LATEST_JSON = "jac_utp_latest.json"

CANDIDATES_HEADER = ["Бренд", "Серия", "№", "Текст УТП", "Брать"]


@dataclass
class UtpCandidate:
    brand: str
    series: str
    text: str


def _write_candidate_rows(rows, xlsx_path: Path, json_path: Path) -> None:
    """Пишет строки (brand, series, text, take) в xlsx + json-бэкап.
    Сортировка по бренду->серии (стабильная — порядок УТП в серии сохраняется);
    нумерация № в пределах серии."""
    from openpyxl import Workbook

    rows = sorted(rows, key=lambda r: (r[0], r[1]))
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "УТП кандидаты"
    ws.append(CANDIDATES_HEADER)
    n_by_series: dict = {}
    for brand, series, text, take in rows:
        key = (brand, series)
        n_by_series[key] = n_by_series.get(key, 0) + 1
        ws.append([brand, series, n_by_series[key], text, take])
    widths = [18, 28, 5, 70, 8]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    wb.save(xlsx_path)

    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps([{"brand": b, "series": s, "text": t} for b, s, t, _ in rows],
                   ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_candidates(cands: List[UtpCandidate], xlsx_path: Path, json_path: Path) -> None:
    """Пишет кандидатов в xlsx (с пустой колонкой «Брать») и json-бэкап."""
    _write_candidate_rows([(c.brand, c.series, c.text, None) for c in cands],
                          xlsx_path, json_path)


def replace_brand_candidates(xlsx_path: Path, json_path: Path, brand: str,
                             new_cands: List[UtpCandidate]) -> int:
    """Заменяет строки одного бренда в файле кандидатов, СОХРАНЯЯ строки и галочки
    остальных брендов. Новые строки бренда добавляются без отметки. Возвращает число
    сохранённых чужих строк."""
    from openpyxl import load_workbook

    kept = []
    if xlsx_path.exists():
        ws = load_workbook(xlsx_path).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5 or row[0] is None:
                continue
            b, s, _num, text, take = row[0], row[1], row[2], row[3], row[4]
            if str(b) != brand:
                kept.append((str(b), str(s), str(text), take))
    rows = kept + [(c.brand, c.series, c.text, None) for c in new_cands]
    _write_candidate_rows(rows, xlsx_path, json_path)
    return len(kept)


def coverage_gaps(jac_series: dict, candidates: List["UtpCandidate"],
                  type_words: tuple) -> dict:
    """Возвращает {бренд: [СЕРИИ_НОРМ без УТП]}, исключая серии-типы.
    jac_series: {бренд: [имена серий JAC]}; candidates: собранные УТП-кандидаты."""
    have = {}
    for c in candidates:
        have.setdefault(c.brand, set()).add(normalize_series(c.series))
    gaps: dict = {}
    for brand, names in jac_series.items():
        brand_have = have.get(brand, set())
        missing = []
        for raw in names:
            norm = normalize_series(raw)
            if not norm or any(w in raw.lower() for w in type_words):
                continue
            if norm not in brand_have and norm not in missing:
                missing.append(norm)
        if missing:
            gaps[brand] = missing
    return gaps


def build_latest_from_xlsx(xlsx_path: Path, out_path: Path) -> dict:
    """Читает отмеченный xlsx, берёт строки с непустой «Брать»,
    группирует в {бренд: {СЕРИЯ_НОРМ: [тексты в порядке файла]}} и пишет json."""
    from openpyxl import load_workbook

    ws = load_workbook(xlsx_path, read_only=True).active
    rows = ws.iter_rows(min_row=2, values_only=True)
    out: dict = {}
    for row in rows:
        if not row or len(row) < 5:
            continue
        brand, series, _num, text, take = row[0], row[1], row[2], row[3], row[4]
        if take is None or str(take).strip() == "" or not text:
            continue
        skey = normalize_series(str(series))
        out.setdefault(str(brand), {}).setdefault(skey, []).append(str(text).strip())

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    return out
